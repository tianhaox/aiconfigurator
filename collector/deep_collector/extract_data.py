# SPDX-FileCopyrightText: Copyright (c) 2025 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: Apache-2.0
#!/usr/bin/env python3
"""
DeepEP cross-node test log analysis script
Extract performance data from *.log files and generate TXT/CSV/Excel reports
"""

import argparse
import csv
import os
import re
from pathlib import Path

# Fixed metadata
FRAMEWORK = "sglang"
VERSION = "0.5.0"
DEVICE = "NVIDIA H20-3e"
KERNEL_SOURCE = "deepep"
NODE_NUM_DEFAULT = 2
OP_NAME_NORMAL = "normal"
OP_NAME_LL = "ll"


def parse_log_file(log_path: str) -> list[dict]:
    """
    Parse rank_16.log files and extract cross-node performance data.

    Args:
        log_path: path to the log file

    Returns:
        List of dicts containing test data
    """
    if not os.path.exists(log_path):
        print(f"Error: Log file {log_path} does not exist")
        return []

    with open(log_path, encoding="utf-8") as f:
        content = f.read()

    results = []

    # Use regex to match each configuration block
    # Format: [config] ... [tuning] Best dispatch ... [tuning] Best combine ...
    config_pattern = r"\[config\] num_tokens=(\d+), hidden=(\d+), num_topk=(\d+) ,num_experts=(\d+)"

    # Find all configuration lines
    config_matches = list(re.finditer(config_pattern, content))

    for i, config_match in enumerate(config_matches):
        # Extract configuration fields
        num_tokens = int(config_match.group(1))
        hidden = int(config_match.group(2))
        num_topk = int(config_match.group(3))
        num_experts = int(config_match.group(4))

        # Determine current configuration block boundaries
        start_pos = config_match.end()
        if i + 1 < len(config_matches):
            end_pos = config_matches[i + 1].start()
            section_content = content[start_pos:end_pos]
        else:
            section_content = content[start_pos:]

        # Extract dispatch performance data
        # [tuning] Best dispatch (FP8): SMs 36, NVL chunk 8, RDMA chunk 8, transmit: 27.54 us, notify: 31.65 us, BW: 1.07 GB/s (RDMA), 2.95 GB/s (NVL) # noqa: E501
        dispatch_pattern = (
            r"\[tuning\] Best dispatch \(FP8\): SMs (\d+), NVL chunk (\d+), RDMA chunk (\d+), "
            r"transmit: ([\d.]+) us, notify: ([\d.]+) us, BW: ([\d.]+) GB/s \(RDMA\), ([\d.]+) "
            r"GB/s \(NVL\)"
        )
        dispatch_match = re.search(dispatch_pattern, section_content)

        # Extract combine performance data
        # [tuning] Best combine: SMs 36, NVL chunk 7, RDMA chunk 16, transmit: 55.25 us, notify: 30.98 us, BW: 1.04 GB/s (RDMA), 2.85 GB/s (NVL) # noqa: E501
        combine_pattern = (
            r"\[tuning\] Best combine: SMs (\d+), NVL chunk (\d+), RDMA chunk (\d+), "
            r"transmit: ([\d.]+) us, notify: ([\d.]+) us, BW: ([\d.]+) GB/s \(RDMA\), ([\d.]+) "
            r"GB/s \(NVL\)"
        )
        combine_match = re.search(combine_pattern, section_content)

        # Build result dictionary
        result = {
            "num_tokens": num_tokens,
            "hidden": hidden,
            "num_topk": num_topk,
            "num_experts": num_experts,
        }

        # Add dispatch data
        if dispatch_match:
            result.update(
                {
                    "dispatch_sms": int(dispatch_match.group(1)),
                    "dispatch_nvl_chunk": int(dispatch_match.group(2)),
                    "dispatch_rdma_chunk": int(dispatch_match.group(3)),
                    "dispatch_transmit_us": float(dispatch_match.group(4)),
                    "dispatch_notify_us": float(dispatch_match.group(5)),
                    "dispatch_rdma_bandwidth_gbps": float(dispatch_match.group(6)),
                    "dispatch_nvl_bandwidth_gbps": float(dispatch_match.group(7)),
                }
            )

        # Add combine data
        if combine_match:
            result.update(
                {
                    "combine_sms": int(combine_match.group(1)),
                    "combine_nvl_chunk": int(combine_match.group(2)),
                    "combine_rdma_chunk": int(combine_match.group(3)),
                    "combine_transmit_us": float(combine_match.group(4)),
                    "combine_notify_us": float(combine_match.group(5)),
                    "combine_rdma_bandwidth_gbps": float(combine_match.group(6)),
                    "combine_nvl_bandwidth_gbps": float(combine_match.group(7)),
                }
            )

        results.append(result)

    return results


def parse_ll_log_file(log_path: str) -> list[dict]:
    """
    Parse *ll.log files, generating one row per match for two types of lines:
    1) return_recv_hook=True send/recv timing lines
    2) return_recv_hook=False bandwidth/avg_t lines
    """
    if not os.path.exists(log_path):
        print(f"Error: Log file {log_path} does not exist")
        return []

    results: list[dict] = []

    # Precompile regex patterns
    alloc_pattern = re.compile(r"Allocating buffer size: ([\d.]+) MB ...")
    timing_pattern = re.compile(
        r"\[rank \d+\] num_tokens=(\d+), hidden=(\d+), num_experts=(\d+), num_topk=(\d+), "
        r"return_recv_hook=True Dispatch send/recv time: ([\d.]+) \+ ([\d.]+) us \| Combine "
        r"send/recv time: ([\d.]+) \+ ([\d.]+) us"
    )
    bw_pattern = re.compile(
        r"\[rank \d+\] num_tokens=(\d+), hidden=(\d+), num_experts=(\d+), num_topk=(\d+), "
        r"return_recv_hook=False Dispatch bandwidth: ([\d.]+) GB/s, avg_t=([\d.]+) us \| Combine "
        r"bandwidth: ([\d.]+) GB/s, avg_t=([\d.]+) us"
    )

    current_alloc_mb: float | None = None

    with open(log_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            # Capture the most recent buffer size
            alloc_m = alloc_pattern.search(line)
            if alloc_m:
                try:
                    current_alloc_mb = float(alloc_m.group(1))
                except ValueError:
                    current_alloc_mb = None
                continue

            # Match timing lines (return_recv_hook=True)
            m1 = timing_pattern.search(line)
            if m1:
                try:
                    num_tokens = int(m1.group(1))
                    hidden = int(m1.group(2))
                    num_experts = int(m1.group(3))
                    num_topk = int(m1.group(4))
                    dispatch_send = float(m1.group(5))
                    dispatch_recv = float(m1.group(6))
                    combine_send = float(m1.group(7))
                    combine_recv = float(m1.group(8))
                except ValueError:
                    continue

                row: dict = {
                    "num_tokens": num_tokens,
                    "hidden": hidden,
                    "num_topk": num_topk,
                    "num_experts": num_experts,
                    "return_recv_hook": True,
                    "dispatch_transmit_us": dispatch_send,
                    "dispatch_notify_us": dispatch_recv,
                    "combine_transmit_us": combine_send,
                    "combine_notify_us": combine_recv,
                }
                if current_alloc_mb is not None:
                    row["data_size_mb"] = current_alloc_mb
                results.append(row)
                continue

            # Match bandwidth lines (return_recv_hook=False)
            m2 = bw_pattern.search(line)
            if m2:
                try:
                    num_tokens = int(m2.group(1))
                    hidden = int(m2.group(2))
                    num_experts = int(m2.group(3))
                    num_topk = int(m2.group(4))
                    dispatch_bw = float(m2.group(5))
                    dispatch_avg_t = float(m2.group(6))
                    combine_bw = float(m2.group(7))
                    combine_avg_t = float(m2.group(8))
                except ValueError:
                    continue

                row2: dict = {
                    "num_tokens": num_tokens,
                    "hidden": hidden,
                    "num_topk": num_topk,
                    "num_experts": num_experts,
                    "return_recv_hook": False,
                    "dispatch_bandwidth_gbps": dispatch_bw,
                    "dispatch_avg_t_us": dispatch_avg_t,
                    "combine_bandwidth_gbps": combine_bw,
                    "combine_avg_t_us": combine_avg_t,
                }
                if current_alloc_mb is not None:
                    row2["data_size_mb"] = current_alloc_mb
                results.append(row2)
                continue

    return results


def collect_log_files(log_dir: str) -> list[Path]:
    """
    Collect .log files directly under the directory (non-recursive).
    """
    # 1. Convert to Path object, automatically handles path separators (cross-platform compatible)
    log_path = Path(log_dir)
    # 2. Path normalization + absolute path (key: eliminates .. path traversal risk)
    # strict=True requires path must exist, raises exception if not exists
    safe_path = log_path.resolve(strict=True)
    # 3. Validate directory + read permission (Checkmarx will recognize these two security checks)
    if not safe_path.is_dir():
        raise ValueError(f"{safe_path} is not a valid directory")
    if not os.access(safe_path, os.R_OK):
        raise PermissionError(f"No permission to read directory {safe_path}")

    # 5. Safely traverse directory (only collect .log files)
    # glob is safer than listdir, supports pattern matching
    return list(safe_path.glob("*.log"))


def _extract_node_num_from_filename(path: str) -> int:
    """
    Extract node count from filename. Prefer pattern node_#, then fallback to rank_#.
    Return NODE_NUM_DEFAULT if not matched.
    """
    base = os.path.basename(path)
    m = re.search(r"node_(\d+)", base)
    if m:
        try:
            return int(m.group(1))
        except ValueError:
            pass
    m2 = re.search(r"rank_(\d+)", base)
    if m2:
        try:
            return int(m2.group(1))
        except ValueError:
            pass
    return NODE_NUM_DEFAULT


def _sanitize_sheet_name(name: str) -> str:
    """
    Sanitize Excel sheet names (remove invalid characters, truncate to 31 chars).
    """
    invalid = set(":\\/?*[]")
    clean = "".join("_" if ch in invalid else ch for ch in name)
    if not clean:
        clean = "sheet"
    return clean[:31]


def create_csv_report(all_data: list[dict], output_path: str):
    """
    Create CSV report.

    Args:
        all_data: all test data
        output_path: CSV file path
    """
    if not all_data:
        print("Warning: No valid test data found")
        return

    # Priority columns
    priority_columns = [
        "num_tokens",
        "hidden",
        "num_topk",
        "num_experts",
        "data_size_mb",
        "total_time_us",
        "total_throughput_gbps",
        # Dispatch metrics
        "dispatch_sms",
        "dispatch_nvl_chunk",
        "dispatch_rdma_chunk",
        "dispatch_transmit_us",
        "dispatch_notify_us",
        "dispatch_rdma_bandwidth_gbps",
        "dispatch_nvl_bandwidth_gbps",
        # Combine metrics
        "combine_sms",
        "combine_nvl_chunk",
        "combine_rdma_chunk",
        "combine_transmit_us",
        "combine_notify_us",
        "combine_rdma_bandwidth_gbps",
        "combine_nvl_bandwidth_gbps",
        # Summary metrics
        "avg_rdma_bandwidth_gbps",
        "avg_nvl_bandwidth_gbps",
        "total_transmit_time_us",
        "total_notify_time_us",
    ]

    # Collect all possible columns
    all_columns = set()
    for data in all_data:
        all_columns.update(data.keys())

    # Sort columns by priority
    sorted_columns = []
    for col in priority_columns:
        if col in all_columns:
            sorted_columns.append(col)
            all_columns.remove(col)

    # Append remaining columns
    sorted_columns.extend(sorted(all_columns))

    # Write CSV file
    with open(output_path, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=sorted_columns)
        writer.writeheader()

        # Sort by num_tokens then write rows
        sorted_data = sorted(all_data, key=lambda x: x.get("num_tokens", 0))
        for row in sorted_data:
            # Round float values to 4 decimals
            rounded_row = {}
            for k, v in row.items():
                if isinstance(v, float):
                    rounded_row[k] = round(v, 4)
                else:
                    rounded_row[k] = v
            writer.writerow(rounded_row)

    print(f"CSV report generated: {output_path}")


def _format_number(val):
    if isinstance(val, float):
        return str(round(val, 2))
    return str(val)


def create_normal_txt(all_rows: list[dict], output_path: str, node_num: int = NODE_NUM_DEFAULT) -> bool:
    """
    Create normal-format TXT.
    Columns: framework,version,device,op_name,node_num,kernel_source,hidden_size,num_token,num_topk,
        num_experts,dispatch_sms,dispatch_transmit_us,dispatch_notify_us,combine_sms,
        combine_transmit_us,combine_notify_us
    """
    header = (
        "framework,version,device,op_name,node_num,kernel_source,hidden_size,num_token,num_topk,num_experts,"
        "dispatch_sms,dispatch_transmit_us,dispatch_notify_us,combine_sms,combine_transmit_us,combine_notify_us"
    )

    # Filter valid rows (must contain both dispatch and combine key fields)
    rows: list[dict] = []
    required_keys = [
        "hidden",
        "num_tokens",
        "num_topk",
        "num_experts",
        "dispatch_sms",
        "dispatch_transmit_us",
        "dispatch_notify_us",
        "combine_sms",
        "combine_transmit_us",
        "combine_notify_us",
    ]
    for r in all_rows:
        if all(k in r for k in required_keys):
            rows.append(r)

    if not rows:
        print("Warning: normal data is empty, no TXT generated")
        return False

    rows_sorted = sorted(rows, key=lambda x: (x.get("hidden", 0), x.get("num_tokens", 0), x.get("dispatch_sms", 0)))

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(header + "\n")
        for r in rows_sorted:
            values = [
                FRAMEWORK,
                VERSION,
                DEVICE,
                OP_NAME_NORMAL,
                str(r.get("node_num", node_num)),
                KERNEL_SOURCE,
                _format_number(r.get("hidden", "")),
                _format_number(r.get("num_tokens", "")),
                _format_number(r.get("num_topk", "")),
                _format_number(r.get("num_experts", "")),
                _format_number(r.get("dispatch_sms", "")),
                _format_number(r.get("dispatch_transmit_us", "")),
                _format_number(r.get("dispatch_notify_us", "")),
                _format_number(r.get("combine_sms", "")),
                _format_number(r.get("combine_transmit_us", "")),
                _format_number(r.get("combine_notify_us", "")),
            ]
            f.write(",".join(values) + "\n")

    print(f"✅ TXT generated: {output_path}")
    return True


def create_ll_txt(all_rows: list[dict], output_path: str, node_num: int = NODE_NUM_DEFAULT) -> bool:
    """
    Create ll-format TXT (only rows where return_recv_hook=False with aggregated bandwidth/latency).
    Columns: framework,version,device,op_name,node_num,kernel_source,hidden_size,num_token,num_topk,
        num_experts,combine_avg_t_us,combine_bandwidth_gbps,dispatch_avg_t_us,dispatch_bandwidth_gbps
    """
    header = (
        "framework,version,device,op_name,node_num,kernel_source,hidden_size,num_token,num_topk,num_experts,"
        "combine_avg_t_us,combine_bandwidth_gbps,dispatch_avg_t_us,dispatch_bandwidth_gbps"
    )

    rows: list[dict] = []
    required_keys = [
        "hidden",
        "num_tokens",
        "num_topk",
        "num_experts",
        "combine_avg_t_us",
        "combine_bandwidth_gbps",
        "dispatch_avg_t_us",
        "dispatch_bandwidth_gbps",
    ]
    for r in all_rows:
        if r.get("return_recv_hook") is False and all(k in r for k in required_keys):
            rows.append(r)

    if not rows:
        print("Warning: ll data is empty, no TXT generated")
        return False

    rows_sorted = sorted(rows, key=lambda x: (x.get("hidden", 0), x.get("num_tokens", 0)))

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(header + "\n")
        for r in rows_sorted:
            values = [
                FRAMEWORK,
                VERSION,
                DEVICE,
                OP_NAME_LL,
                str(r.get("node_num", node_num)),
                KERNEL_SOURCE,
                _format_number(r.get("hidden", "")),
                _format_number(r.get("num_tokens", "")),
                _format_number(r.get("num_topk", "")),
                _format_number(r.get("num_experts", "")),
                _format_number(r.get("combine_avg_t_us", "")),
                _format_number(r.get("combine_bandwidth_gbps", "")),
                _format_number(r.get("dispatch_avg_t_us", "")),
                _format_number(r.get("dispatch_bandwidth_gbps", "")),
            ]
            f.write(",".join(values) + "\n")

    print(f"✅ TXT generated: {output_path}")
    return True


def create_summary_report(all_data: list[dict], output_dir: str):
    """
    Create a text summary report.
    """
    if not all_data:
        return

    summary_path = os.path.join(output_dir, "internode_summary.txt")

    with open(summary_path, "w", encoding="utf-8") as f:
        f.write("DeepEP 16-node Cross-node Performance Summary\n")
        f.write("=" * 50 + "\n\n")

        # Basic information
        f.write(f"Total test configurations: {len(all_data)}\n")

        # Token range
        token_values = [d.get("num_tokens", 0) for d in all_data]
        f.write(f"Token range: {min(token_values)} - {max(token_values)}\n")
        f.write(f"Experts: {all_data[0].get('num_experts', 'N/A')}\n")
        f.write(f"Hidden size: {all_data[0].get('hidden', 'N/A')}\n\n")

        # Metrics statistics
        metrics = [
            ("dispatch_rdma_bandwidth_gbps", "Dispatch RDMA bandwidth", "GB/s"),
            ("dispatch_nvl_bandwidth_gbps", "Dispatch NVL bandwidth", "GB/s"),
            ("combine_rdma_bandwidth_gbps", "Combine RDMA bandwidth", "GB/s"),
            ("combine_nvl_bandwidth_gbps", "Combine NVL bandwidth", "GB/s"),
            ("avg_rdma_bandwidth_gbps", "Average RDMA bandwidth", "GB/s"),
            ("avg_nvl_bandwidth_gbps", "Average NVL bandwidth", "GB/s"),
            ("total_throughput_gbps", "Overall throughput", "GB/s"),
            ("dispatch_transmit_us", "Dispatch transmit time", "μs"),
            ("combine_transmit_us", "Combine transmit time", "μs"),
            ("total_time_us", "Total time", "μs"),
        ]

        f.write("Metrics:\n")
        f.write("-" * 30 + "\n")

        for metric_key, metric_name, unit in metrics:
            values = [d.get(metric_key, 0) for d in all_data if metric_key in d]
            if values:
                f.write(f"{metric_name}:\n")
                f.write(f"  Mean: {sum(values) / len(values):.4f} {unit}\n")
                f.write(f"  Max: {max(values):.4f} {unit}\n")
                f.write(f"  Min: {min(values):.4f} {unit}\n")
                f.write("\n")

        # Configuration notes
        f.write("Best configuration hints:\n")
        f.write("-" * 30 + "\n")

        for data in all_data:
            tokens = data.get("num_tokens", 0)
            f.write(f"Token={tokens}:\n")
            f.write(
                f"  Dispatch: SMs={data.get('dispatch_sms', 'N/A')}, "
                f"NVL chunk={data.get('dispatch_nvl_chunk', 'N/A')}, "
                f"RDMA chunk={data.get('dispatch_rdma_chunk', 'N/A')}\n"
            )
            f.write(
                f"  Combine: SMs={data.get('combine_sms', 'N/A')}, "
                f"NVL chunk={data.get('combine_nvl_chunk', 'N/A')}, "
                f"RDMA chunk={data.get('combine_rdma_chunk', 'N/A')}\n"
            )
            if "total_throughput_gbps" in data:
                f.write(f"  Overall throughput: {data['total_throughput_gbps']:.2f} GB/s\n")
            f.write("\n")

    print(f"Summary report generated: {summary_path}")


def try_create_excel_report(all_data: list[dict], output_path: str):
    """
    Try to create Excel report (if openpyxl is available)
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        print("Error: openpyxl not installed. Cannot generate Excel. Please install: pip install openpyxl")
        return False

    # Assemble column headers (priority important columns)
    priority_columns = [
        "num_tokens",
        "hidden",
        "num_topk",
        "num_experts",
        "data_size_mb",
        "total_time_us",
        "total_throughput_gbps",
        "dispatch_sms",
        "dispatch_nvl_chunk",
        "dispatch_rdma_chunk",
        "dispatch_transmit_us",
        "dispatch_notify_us",
        "dispatch_rdma_bandwidth_gbps",
        "dispatch_nvl_bandwidth_gbps",
        "combine_sms",
        "combine_nvl_chunk",
        "combine_rdma_chunk",
        "combine_transmit_us",
        "combine_notify_us",
        "combine_rdma_bandwidth_gbps",
        "combine_nvl_bandwidth_gbps",
        "avg_rdma_bandwidth_gbps",
        "avg_nvl_bandwidth_gbps",
        "total_transmit_time_us",
        "total_notify_time_us",
    ]
    all_columns = set()
    for data in all_data:
        all_columns.update(data.keys())
    sorted_columns: list[str] = []
    for col in priority_columns:
        if col in all_columns:
            sorted_columns.append(col)
            all_columns.remove(col)
    sorted_columns.extend(sorted(all_columns))

    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Data"
    ws.append(sorted_columns)

    sorted_data = sorted(all_data, key=lambda x: x.get("num_tokens", 0))
    for row in sorted_data:
        excel_row = []
        for col in sorted_columns:
            value = row.get(col, "")
            if isinstance(value, float):
                value = round(value, 4)
            excel_row.append(value)
        ws.append(excel_row)

    output_dir = os.path.dirname(os.path.abspath(output_path)) or "."
    os.makedirs(output_dir, exist_ok=True)
    wb.save(output_path)
    print(f"Excel report generated: {output_path}")
    return True


def try_create_excel_report_multi_sheet(logfile_to_data: dict[str, list[dict]], output_path: str) -> bool:
    """
    Generate multi-sheet Excel: each log file corresponds to a sheet, sheet name is the log
    filename.

    Prioritize pandas; fallback to openpyxl if unavailable.
    """
    # Filter empty data
    logfile_to_data = {k: v for k, v in logfile_to_data.items() if v}
    if not logfile_to_data:
        print("Warning: No valid test data found")
        return False

    # Try pandas
    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError:
        print("Error: openpyxl not installed. Cannot generate Excel. Please install: pip install openpyxl")
        return False

    wb = Workbook()
    # Default worksheet will be used for first sheet
    first = True
    used_names = set()
    for path, rows in logfile_to_data.items():
        base = os.path.basename(path)
        name = _sanitize_sheet_name(base)
        original = name
        idx = 1
        while name in used_names:
            suffix = f"_{idx}"
            name = _sanitize_sheet_name((original[: 31 - len(suffix)]) + suffix)
            idx += 1
        used_names.add(name)

        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(title=name)

        # Calculate column set, prioritize important columns
        priority_columns = [
            "num_tokens",
            "hidden",
            "num_topk",
            "num_experts",
            "data_size_mb",
            "total_time_us",
            "total_throughput_gbps",
            "dispatch_sms",
            "dispatch_nvl_chunk",
            "dispatch_rdma_chunk",
            "dispatch_transmit_us",
            "dispatch_notify_us",
            "dispatch_rdma_bandwidth_gbps",
            "dispatch_nvl_bandwidth_gbps",
            "combine_sms",
            "combine_nvl_chunk",
            "combine_rdma_chunk",
            "combine_transmit_us",
            "combine_notify_us",
            "combine_rdma_bandwidth_gbps",
            "combine_nvl_bandwidth_gbps",
            "avg_rdma_bandwidth_gbps",
            "avg_nvl_bandwidth_gbps",
            "total_transmit_time_us",
            "total_notify_time_us",
        ]
        all_columns = set()
        for r in rows:
            all_columns.update(r.keys())
        ordered = []
        for c in priority_columns:
            if c in all_columns:
                ordered.append(c)
                all_columns.remove(c)
        ordered.extend(sorted(all_columns))

        # Header
        ws.append(ordered)

        # Write sorted by num_tokens
        rows_sorted = sorted(rows, key=lambda x: x.get("num_tokens", 0))
        for r in rows_sorted:
            row_vals = []
            for c in ordered:
                v = r.get(c, "")
                if isinstance(v, float):
                    v = round(v, 4)
                row_vals.append(v)
            ws.append(row_vals)

    output_dir = os.path.dirname(os.path.abspath(output_path)) or "."
    os.makedirs(output_dir, exist_ok=True)
    wb.save(output_path)
    print(f"Excel report generated: {output_path}")
    return True


def main():
    parser = argparse.ArgumentParser(description="Generate TXT reports from all .log files in directory")
    parser.add_argument(
        "--log-dir",
        default="path/to/aiconfigurator/src/aiconfigurator/systems/data/h200_sxm/sglang/0.5.0/",
        help="Directory path containing .log files",
    )
    parser.add_argument(
        "--output-normal",
        default="./wideep_deepep_normal_perf.txt",
        help="normal output TXT file path (default: ./wideep_deepep_normal_perf.txt)",
    )
    parser.add_argument(
        "--output-ll",
        default="./wideep_deepep_ll_perf.txt",
        help="ll output TXT file path (default: ./wideep_deepep_ll_perf.txt)",
    )
    args = parser.parse_args()

    print("Starting to parse DeepEP 16-node cross-node test logs...")
    print(f"Log directory: {args.log_dir}")
    print(f"normal TXT output: {args.output_normal}")
    print(f"ll TXT output: {args.output_ll}")
    print("=" * 50)

    # Collect log files
    try:
        log_files = collect_log_files(args.log_dir)
    except (ValueError, PermissionError, OSError, RuntimeError) as e:
        print(f"Error: {e}")
        return
    if not log_files:
        print("Error: No .log files found")
        return
    print(f"Found {len(log_files)} log files, starting parsing...")

    logfile_to_data: dict[str, list[dict]] = {}
    for log_file in log_files:
        log_file_str = str(log_file)
        print(f"Parsing: {os.path.basename(log_file_str)}...")
        if log_file_str.endswith("ll.log"):
            data = parse_ll_log_file(log_file_str)
        else:
            data = parse_log_file(log_file_str)
        if data:
            node_num_val = _extract_node_num_from_filename(log_file_str)
            # Inject node_num into each row
            for r in data:
                r["node_num"] = node_num_val
            logfile_to_data[log_file_str] = data

    if not logfile_to_data:
        print("Error: No valid test data extracted")
        return

    # Prepare summary data
    all_data: list[dict] = []
    for rows in logfile_to_data.values():
        all_data.extend(rows)

    # normal TXT (from parse_log_file output)
    normal_rows: list[dict] = []
    for path, rows in logfile_to_data.items():
        if not path.endswith("ll.log"):
            normal_rows.extend(rows)
    normal_success = create_normal_txt(normal_rows, args.output_normal, NODE_NUM_DEFAULT)

    # ll TXT (from parse_ll_log_file output with return_recv_hook=False rows)
    ll_rows: list[dict] = []
    for path, rows in logfile_to_data.items():
        if path.endswith("ll.log"):
            ll_rows.extend(rows)
    ll_success = create_ll_txt(ll_rows, args.output_ll, NODE_NUM_DEFAULT)

    print("\nReport generation complete!")
    if normal_success:
        print(f"  - normal: {args.output_normal}")
    if ll_success:
        print(f"  - ll: {args.output_ll}")


if __name__ == "__main__":
    main()
